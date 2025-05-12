# email_downloader.py

import imaplib
import email
from email.header import decode_header
from email.utils import parseaddr
import os
from datetime import datetime
from openpyxl import load_workbook
from config import EMAIL, SENHA, IMAP_SERVER, PASTA_DESTINO, LOG_PATH, WHITELISTED_SENDERS
from pdf_to_excel import main as gerar_excels_dos_pdfs

EXTENSOES_PERMITIDAS = ('.pdf',)
os.makedirs(PASTA_DESTINO, exist_ok=True)
os.makedirs(os.path.dirname(LOG_PATH), exist_ok=True)

def log(mensagem):
    with open(LOG_PATH, "a", encoding="utf-8") as f:
        f.write(f"[{datetime.now()}] {mensagem}\n")

def conectar_email():
    mail = imaplib.IMAP4_SSL(IMAP_SERVER)
    mail.login(EMAIL, SENHA)
    mail.select("inbox")
    return mail

def buscar_emails(mail, palavra_chave):
    status, mensagens = mail.search(None, f'(UNSEEN SUBJECT "{palavra_chave}")')
    return mensagens[0].split()

def abrir_excel(path):
    try:
        print(f"📊 Abrindo e lendo Excel: {path}")
        wb = load_workbook(filename=path)
        sheet = wb.active
        for row in sheet.iter_rows(values_only=True):
            print(row)
    except Exception as e:
        print(f"⚠️ Erro ao ler Excel {path}: {e}")

def baixar_anexos(mail, mensagens):
    for num in mensagens:
        status, dados = mail.fetch(num, '(RFC822)')
        raw_email = dados[0][1]
        msg = email.message_from_bytes(raw_email)

        assunto, codif = decode_header(msg.get("Subject", "Sem assunto"))[0]
        assunto = assunto.decode(codif or "utf-8") if isinstance(assunto, bytes) else assunto

        remetente = msg.get("From", "Desconhecido")
        remetente_email = parseaddr(remetente)[1]

        if remetente_email not in WHITELISTED_SENDERS:
            log(f"⛔ Remetente não autorizado: {remetente_email}")
            continue

        log(f"📥 E-mail de: {remetente_email} | Assunto: {assunto}")

        for parte in msg.walk():
            if parte.get_content_maintype() == 'multipart':
                continue
            if parte.get("Content-Disposition") is None:
                continue

            nome_arquivo = parte.get_filename()
            if not nome_arquivo:
                continue

            nome_arquivo, codif = decode_header(nome_arquivo)[0]
            if isinstance(nome_arquivo, bytes):
                nome_arquivo = nome_arquivo.decode(codif or "utf-8")

            if not nome_arquivo.lower().endswith(EXTENSOES_PERMITIDAS):
                log(f"⚠️ Arquivo ignorado (não-PDF): {nome_arquivo}")
                continue

            base, ext = os.path.splitext(nome_arquivo)
            caminho = os.path.join(PASTA_DESTINO, nome_arquivo)
            contador = 1
            while os.path.exists(caminho):
                caminho = os.path.join(PASTA_DESTINO, f"{base}_{contador}{ext}")
                contador += 1

            with open(caminho, "wb") as f:
                f.write(parte.get_payload(decode=True))

            log(f"📎 Anexo PDF salvo: {caminho}")

    mail.logout()
    gerar_excels_dos_pdfs()

# Execução principal
if __name__ == "__main__":
    try:
        mail = conectar_email()
        baixar_anexos(mail)
    except Exception as e:
        log(f"❌ Erro durante execução: {e}")
